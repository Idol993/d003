#!/usr/bin/env python3
"""Banking Core Transaction System Release & Compliance Rollback Automation Manager"""

import os
import sys
import json
import csv
import hashlib
import logging
import sqlite3
import threading
import time
import zipfile
import argparse
import random
import traceback
from datetime import datetime, timedelta
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Optional, List, Dict, Any, Callable
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

BASE_DIR = Path(__file__).parent.resolve()
DB_PATH = BASE_DIR / "bank_core_release.db"
EXPORTS_DIR = BASE_DIR / "exports"
AUDIT_LOGS_DIR = BASE_DIR / "audit_logs"
ARCHIVES_DIR = BASE_DIR / "archives"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("BankCoreRelease")


# ─── Enums ────────────────────────────────────────────────────────────────────

class RiskLevel(Enum):
    NORMAL_DAYTIME = "日间非交易"
    NIGHTTIME_BATCH = "夜间批处理"
    EMERGENCY_FAULT = "紧急故障"


class ReleaseStatus(Enum):
    SUBMITTED = "submitted"
    PRECHECK_PASSED = "precheck_passed"
    PRECHECK_FAILED = "precheck_failed"
    PENDING_APPROVAL = "pending_approval"
    APPROVED = "approved"
    APPROVED_DEPLOY_FAILED = "approved_deploy_failed"
    REJECTED = "rejected"
    DEPLOYING = "deploying"
    DEPLOYED = "deployed"
    ROLLED_BACK = "rolled_back"
    FAILED = "failed"


class GrayscalePhase(Enum):
    PHASE_5 = "5%"
    PHASE_20 = "20%"
    PHASE_50 = "50%"
    PHASE_100 = "100%"
    COMPLETED = "completed"


class MonitorStatus(Enum):
    RUNNING = "running"
    STOPPED = "stopped"
    ANOMALY = "anomaly"


class ApprovalStatus(Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"


class DrillStatus(Enum):
    CREATED = "created"
    IN_PROGRESS = "in_progress"
    COMPLETED = "completed"
    ARCHIVED = "archived"


# ─── Constants ────────────────────────────────────────────────────────────────

APPROVAL_MATRIX = {
    RiskLevel.NORMAL_DAYTIME.value: {
        "nodes": ["总行科技", "总行业务"],
        "timeout_hours": 24,
    },
    RiskLevel.NIGHTTIME_BATCH.value: {
        "nodes": ["总行科技", "总行业务", "总行合规"],
        "timeout_hours": 8,
    },
    RiskLevel.EMERGENCY_FAULT.value: {
        "nodes": ["总行科技", "总行业务", "总行合规"],
        "timeout_hours": 2,
    },
}

GRAYSCALE_PHASES = [
    (GrayscalePhase.PHASE_5, 5, 15),
    (GrayscalePhase.PHASE_20, 20, 15),
    (GrayscalePhase.PHASE_50, 50, 15),
    (GrayscalePhase.PHASE_100, 100, 10),
]

MONITOR_THRESHOLDS = {
    "transaction_success_rate": 99.9,
    "accounting_delay_ms": 500,
    "fund_settlement_anomaly": 0,
}

DRILL_STEPS = [
    "启动确认",
    "通知相关方",
    "数据备份验证",
    "切换执行",
    "功能验证",
    "性能验证",
    "回退演练",
    "总结归档",
]

SAMPLE_BRANCHES = [
    {"code": "BJ001", "name": "北京总行营业部", "region": "华北"},
    {"code": "BJ002", "name": "北京中关村支行", "region": "华北"},
    {"code": "SH001", "name": "上海陆家嘴支行", "region": "华东"},
    {"code": "SH002", "name": "上海外滩支行", "region": "华东"},
    {"code": "GZ001", "name": "广州天河支行", "region": "华南"},
    {"code": "GZ002", "name": "广州珠江新城支行", "region": "华南"},
    {"code": "SZ001", "name": "深圳福田支行", "region": "华南"},
    {"code": "SZ002", "name": "深圳南山支行", "region": "华南"},
    {"code": "CD001", "name": "成都锦江支行", "region": "西南"},
    {"code": "CD002", "name": "成都高新支行", "region": "西南"},
    {"code": "CQ001", "name": "重庆渝中支行", "region": "西南"},
    {"code": "WH001", "name": "武汉江汉支行", "region": "华中"},
    {"code": "NJ001", "name": "南京新街口支行", "region": "华东"},
    {"code": "HZ001", "name": "杭州武林支行", "region": "华东"},
    {"code": "TJ001", "name": "天津和平支行", "region": "华北"},
    {"code": "XA001", "name": "西安钟楼支行", "region": "西北"},
    {"code": "DL001", "name": "大连中山支行", "region": "东北"},
    {"code": "SY001", "name": "沈阳和平支行", "region": "东北"},
    {"code": "CS001", "name": "长沙芙蓉支行", "region": "华中"},
    {"code": "KM001", "name": "昆明五华支行", "region": "西南"},
]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _gen_id(prefix: str = "") -> str:
    ts = datetime.now().strftime("%Y%m%d%H%M%S%f")
    return f"{prefix}{ts}{random.randint(100, 999)}"


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ─── Database Layer ───────────────────────────────────────────────────────────

class Database:
    _lock = threading.Lock()
    _local = threading.local()

    @staticmethod
    def get_conn() -> sqlite3.Connection:
        if not hasattr(Database._local, "conn") or Database._local.conn is None:
            conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.row_factory = sqlite3.Row
            Database._local.conn = conn
        return Database._local.conn

    @staticmethod
    def execute(sql: str, params: tuple = ()) -> sqlite3.Cursor:
        with Database._lock:
            conn = Database.get_conn()
            cur = conn.execute(sql, params)
            conn.commit()
            return cur

    @staticmethod
    def query(sql: str, params: tuple = ()) -> List[sqlite3.Row]:
        with Database._lock:
            conn = Database.get_conn()
            cur = conn.execute(sql, params)
            return cur.fetchall()

    @staticmethod
    def query_one(sql: str, params: tuple = ()) -> Optional[sqlite3.Row]:
        with Database._lock:
            conn = Database.get_conn()
            cur = conn.execute(sql, params)
            return cur.fetchone()


def init_db():
    for d in [EXPORTS_DIR, AUDIT_LOGS_DIR, ARCHIVES_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    conn = Database.get_conn()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS release_applications (
        id TEXT PRIMARY KEY,
        title TEXT NOT NULL,
        version TEXT NOT NULL,
        module TEXT NOT NULL,
        applicant TEXT NOT NULL,
        risk_level TEXT NOT NULL,
        status TEXT DEFAULT 'submitted',
        branch_code TEXT DEFAULT '',
        description TEXT DEFAULT '',
        previous_stable_version TEXT DEFAULT '',
        deploy_failure_reason TEXT DEFAULT '',
        created_at TEXT,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS approval_nodes (
        id TEXT PRIMARY KEY,
        release_id TEXT NOT NULL,
        role TEXT NOT NULL,
        approver TEXT DEFAULT '',
        status TEXT DEFAULT 'pending',
        comment TEXT DEFAULT '',
        approved_at TEXT,
        created_at TEXT,
        FOREIGN KEY (release_id) REFERENCES release_applications(id)
    );
    CREATE TABLE IF NOT EXISTS rollback_records (
        id TEXT PRIMARY KEY,
        release_id TEXT NOT NULL,
        reason TEXT,
        previous_version TEXT,
        affected_accounts INTEGER DEFAULT 0,
        regulatory_explanation TEXT,
        root_cause TEXT,
        status TEXT DEFAULT 'completed',
        created_at TEXT,
        notified_stakeholders TEXT,
        FOREIGN KEY (release_id) REFERENCES release_applications(id)
    );
    CREATE TABLE IF NOT EXISTS drill_records (
        id TEXT PRIMARY KEY,
        title TEXT NOT NULL,
        scenario TEXT,
        status TEXT DEFAULT 'created',
        plan TEXT DEFAULT '',
        steps TEXT DEFAULT '',
        evidence TEXT DEFAULT '',
        created_at TEXT,
        completed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS monitor_snapshots (
        id TEXT PRIMARY KEY,
        release_id TEXT NOT NULL,
        transaction_success_rate REAL,
        accounting_delay_ms REAL,
        fund_settlement_anomaly INTEGER DEFAULT 0,
        status TEXT DEFAULT 'running',
        timestamp TEXT
    );
    CREATE TABLE IF NOT EXISTS audit_log_chain (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp TEXT NOT NULL,
        action TEXT NOT NULL,
        entity_type TEXT,
        entity_id TEXT,
        details TEXT DEFAULT '',
        prev_hash TEXT NOT NULL,
        current_hash TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS stable_versions (
        id TEXT PRIMARY KEY,
        version TEXT NOT NULL,
        module TEXT NOT NULL,
        marked_at TEXT,
        release_id TEXT
    );
    CREATE TABLE IF NOT EXISTS branch_registry (
        code TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        region TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS grayscale_state (
        release_id TEXT PRIMARY KEY,
        current_phase TEXT DEFAULT '',
        phase_index INTEGER DEFAULT 0,
        branches_deployed TEXT DEFAULT '',
        started_at TEXT,
        FOREIGN KEY (release_id) REFERENCES release_applications(id)
    );
    CREATE TABLE IF NOT EXISTS monitor_runs (
        id TEXT PRIMARY KEY,
        release_id TEXT NOT NULL,
        started_at TEXT,
        stopped_at TEXT,
        auto_rollback_enabled INTEGER DEFAULT 0,
        mode TEXT DEFAULT 'thread',
        FOREIGN KEY (release_id) REFERENCES release_applications(id)
    );
    """)
    conn.commit()

    try:
        cols = [r["name"] for r in Database.query("PRAGMA table_info(release_applications)")]
        if "previous_stable_version" not in cols:
            Database.execute("ALTER TABLE release_applications ADD COLUMN previous_stable_version TEXT DEFAULT ''")
        if "deploy_failure_reason" not in cols:
            Database.execute("ALTER TABLE release_applications ADD COLUMN deploy_failure_reason TEXT DEFAULT ''")
    except Exception:
        pass

    rows = Database.query("SELECT COUNT(*) AS cnt FROM branch_registry")
    if rows[0]["cnt"] == 0:
        for b in SAMPLE_BRANCHES:
            Database.execute(
                "INSERT OR IGNORE INTO branch_registry (code, name, region) VALUES (?, ?, ?)",
                (b["code"], b["name"], b["region"]),
            )

    rows = Database.query("SELECT COUNT(*) AS cnt FROM audit_log_chain")
    if rows[0]["cnt"] == 0:
        ts = _now()
        genesis_raw = f"GENESIS|{ts}|CHAIN_INIT|SYSTEM|0|Audit chain initialized"
        genesis_hash = hashlib.sha256(genesis_raw.encode()).hexdigest()
        Database.execute(
            "INSERT INTO audit_log_chain (timestamp, action, entity_type, entity_id, details, prev_hash, current_hash) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ts, "CHAIN_INIT", "SYSTEM", "0", "Audit chain initialized", "GENESIS", genesis_hash),
        )

    rows = Database.query("SELECT COUNT(*) AS cnt FROM stable_versions")
    if rows[0]["cnt"] == 0:
        Database.execute(
            "INSERT INTO stable_versions (id, version, module, marked_at, release_id) VALUES (?, ?, ?, ?, ?)",
            ("stable_001", "3.2.1", "core", "2025-01-01 00:00:00", "legacy"),
        )

    logger.info("Database initialized with WAL mode")


# ─── 2. AuditLogger ──────────────────────────────────────────────────────────

class AuditLogger:
    def __init__(self):
        self._lock = threading.Lock()

    def _get_last_hash(self) -> str:
        row = Database.query_one(
            "SELECT current_hash FROM audit_log_chain ORDER BY id DESC LIMIT 1"
        )
        return row["current_hash"] if row else "GENESIS"

    def log(self, action: str, entity_type: str = "", entity_id: str = "",
            details: str = ""):
        with self._lock:
            prev_hash = self._get_last_hash()
            ts = _now()
            raw = f"{prev_hash}|{ts}|{action}|{entity_type}|{entity_id}|{details}"
            current_hash = hashlib.sha256(raw.encode()).hexdigest()
            Database.execute(
                "INSERT INTO audit_log_chain (timestamp, action, entity_type, entity_id, details, prev_hash, current_hash) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (ts, action, entity_type, entity_id, details, prev_hash, current_hash),
            )
            self._write_to_daily_log(ts, action, entity_type, entity_id, details, current_hash)
        logger.info("Audit: %s [%s/%s] %s", action, entity_type, entity_id, details[:80])

    def _write_to_daily_log(self, ts, action, entity_type, entity_id, details, current_hash):
        date_str = datetime.now().strftime("%Y-%m-%d")
        log_file = AUDIT_LOGS_DIR / f"audit_{date_str}.log"
        line = f"[{ts}] {action} | {entity_type} | {entity_id} | {details} | hash={current_hash}\n"
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(line)

    def verify_integrity(self) -> Dict[str, Any]:
        rows = Database.query(
            "SELECT id, timestamp, action, entity_type, entity_id, details, prev_hash, current_hash "
            "FROM audit_log_chain ORDER BY id ASC"
        )
        if not rows:
            return {"valid": True, "entries": 0, "errors": []}
        errors = []
        prev_hash = "GENESIS"
        for row in rows:
            if row["prev_hash"] != prev_hash:
                errors.append(
                    f"Chain break at id={row['id']}: expected prev_hash={prev_hash}, got={row['prev_hash']}"
                )
            raw = (
                f"{row['prev_hash']}|{row['timestamp']}|{row['action']}|"
                f"{row['entity_type']}|{row['entity_id']}|{row['details']}"
            )
            expected_hash = hashlib.sha256(raw.encode()).hexdigest()
            if row["current_hash"] != expected_hash:
                errors.append(
                    f"Hash mismatch at id={row['id']}: expected={expected_hash}, got={row['current_hash']}"
                )
            prev_hash = row["current_hash"]
        return {"valid": len(errors) == 0, "entries": len(rows), "errors": errors}

    def query(self, action: str = None, entity_type: str = None, entity_id: str = None,
              start_time: str = None, end_time: str = None, limit: int = 100) -> List[Dict]:
        conditions = []
        params: list = []
        if action:
            conditions.append("action = ?")
            params.append(action)
        if entity_type:
            conditions.append("entity_type = ?")
            params.append(entity_type)
        if entity_id:
            conditions.append("entity_id = ?")
            params.append(entity_id)
        if start_time:
            conditions.append("timestamp >= ?")
            params.append(start_time)
        if end_time:
            conditions.append("timestamp <= ?")
            params.append(end_time)
        where = " AND ".join(conditions) if conditions else "1=1"
        sql = f"SELECT * FROM audit_log_chain WHERE {where} ORDER BY id DESC LIMIT ?"
        params.append(limit)
        rows = Database.query(sql, tuple(params))
        return [dict(r) for r in rows]


# ─── 3. PreCheckEngine ───────────────────────────────────────────────────────

class PreCheckEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def run_checks(self, release_id: str, version: str, module: str) -> Dict[str, Any]:
        results = {
            "regulatory_filing": self._check_regulatory_filing(release_id),
            "stress_test": self._check_stress_test(release_id),
            "account_consistency": self._check_account_consistency(release_id),
            "disaster_recovery": self._check_disaster_recovery(release_id),
        }
        passed = all(r["passed"] for r in results.values())
        status = "passed" if passed else "failed"
        self.audit.log(
            "PRECHECK", "ReleaseApplication", release_id,
            json.dumps({"status": status, "results": results}, ensure_ascii=False),
        )
        return {"release_id": release_id, "passed": passed, "results": results}

    def _check_regulatory_filing(self, release_id: str) -> Dict:
        if os.environ.get("skip_filing", "").lower() == "true":
            return {"passed": False, "message": "监管报备未完成（模拟跳过）", "detail": "skip_filing=true"}
        return {"passed": True, "message": "监管报备已完成", "detail": "银保监会报备编号：CBIRC-2026-AUTO"}

    def _check_stress_test(self, release_id: str) -> Dict:
        if os.environ.get("skip_stress", "").lower() == "true":
            rate = 99.5
            return {
                "passed": False,
                "message": f"压力测试通过率{rate}%，未达99.9%阈值（模拟）",
                "detail": f"rate={rate}%",
            }
        rate = 99.95
        return {"passed": rate >= 99.9, "message": f"压力测试通过率{rate}%", "detail": f"rate={rate}%"}

    def _check_account_consistency(self, release_id: str) -> Dict:
        if os.environ.get("skip_consistency", "").lower() == "true":
            return {"passed": False, "message": "账务一致性校验失败（模拟）", "detail": "consistency=99.8%"}
        return {"passed": True, "message": "账务一致性校验100%通过", "detail": "consistency=100%"}

    def _check_disaster_recovery(self, release_id: str) -> Dict:
        if os.environ.get("skip_dr", "").lower() == "true":
            return {"passed": False, "message": "灾备切换不可用（模拟）", "detail": "dr_available=false"}
        return {"passed": True, "message": "灾备切换验证通过", "detail": "RTO=30s, RPO=0"}


# ─── 4. ApprovalEngine ───────────────────────────────────────────────────────

class ApprovalEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def create_workflow(self, release_id: str, risk_level: str) -> List[Dict]:
        matrix = APPROVAL_MATRIX.get(risk_level, APPROVAL_MATRIX[RiskLevel.NORMAL_DAYTIME.value])
        nodes = []
        for i, role in enumerate(matrix["nodes"]):
            node_id = _gen_id("apr_")
            Database.execute(
                "INSERT INTO approval_nodes (id, release_id, role, status, created_at) VALUES (?, ?, ?, ?, ?)",
                (node_id, release_id, role, ApprovalStatus.PENDING.value, _now()),
            )
            nodes.append({"id": node_id, "role": role, "status": ApprovalStatus.PENDING.value, "order": i})
        self.audit.log(
            "APPROVAL_WORKFLOW_CREATED", "ReleaseApplication", release_id,
            json.dumps({
                "risk_level": risk_level,
                "nodes": [n["role"] for n in nodes],
                "timeout_hours": matrix["timeout_hours"],
            }, ensure_ascii=False),
        )
        return nodes

    def get_pending_nodes(self, release_id: str) -> List[Dict]:
        rows = Database.query(
            "SELECT * FROM approval_nodes WHERE release_id = ? AND status = ? ORDER BY id",
            (release_id, ApprovalStatus.PENDING.value),
        )
        return [dict(r) for r in rows]

    def approve_node(self, release_id: str, role: str, approver: str,
                     comment: str = "") -> Dict:
        row = Database.query_one(
            "SELECT * FROM approval_nodes WHERE release_id = ? AND role = ? AND status = ?",
            (release_id, role, ApprovalStatus.PENDING.value),
        )
        if not row:
            return {"success": False, "message": f"未找到待审批节点: {role}"}
        Database.execute(
            "UPDATE approval_nodes SET status = ?, approver = ?, comment = ?, approved_at = ? WHERE id = ?",
            (ApprovalStatus.APPROVED.value, approver, comment, _now(), row["id"]),
        )
        self.audit.log(
            "APPROVE", "ApprovalNode", row["id"],
            json.dumps({"role": role, "approver": approver, "comment": comment}, ensure_ascii=False),
        )
        all_approved = self._all_approved(release_id)
        return {"success": True, "role": role, "approver": approver, "all_approved": all_approved}

    def reject_node(self, release_id: str, role: str, approver: str,
                    comment: str = "") -> Dict:
        row = Database.query_one(
            "SELECT * FROM approval_nodes WHERE release_id = ? AND role = ? AND status = ?",
            (release_id, role, ApprovalStatus.PENDING.value),
        )
        if not row:
            return {"success": False, "message": f"未找到待审批节点: {role}"}
        Database.execute(
            "UPDATE approval_nodes SET status = ?, approver = ?, comment = ?, approved_at = ? WHERE id = ?",
            (ApprovalStatus.REJECTED.value, approver, comment, _now(), row["id"]),
        )
        Database.execute(
            "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
            (ReleaseStatus.REJECTED.value, _now(), release_id),
        )
        self.audit.log(
            "REJECT", "ApprovalNode", row["id"],
            json.dumps({"role": role, "approver": approver, "comment": comment}, ensure_ascii=False),
        )
        return {"success": True, "role": role, "approver": approver, "rejected": True}

    def _all_approved(self, release_id: str) -> bool:
        rows = Database.query(
            "SELECT * FROM approval_nodes WHERE release_id = ?", (release_id,)
        )
        return all(r["status"] == ApprovalStatus.APPROVED.value for r in rows)

    def get_timeout(self, risk_level: str) -> int:
        matrix = APPROVAL_MATRIX.get(risk_level, APPROVAL_MATRIX[RiskLevel.NORMAL_DAYTIME.value])
        return matrix["timeout_hours"]


# ─── 5. GrayscaleDeployEngine ────────────────────────────────────────────────

class GrayscaleDeployEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def init_deployment(self, release_id: str) -> Dict:
        all_branches = [dict(r) for r in Database.query("SELECT * FROM branch_registry")]
        Database.execute(
            "INSERT OR REPLACE INTO grayscale_state (release_id, current_phase, phase_index, branches_deployed, started_at) VALUES (?, ?, ?, ?, ?)",
            (release_id, GrayscalePhase.PHASE_5.value, 0, "[]", _now()),
        )
        Database.execute(
            "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
            (ReleaseStatus.DEPLOYING.value, _now(), release_id),
        )
        self.audit.log("GRAYSCALE_INIT", "ReleaseApplication", release_id, "灰度发布初始化")
        return self._phase_info(release_id, 0, all_branches)

    def advance_phase(self, release_id: str) -> Dict:
        state = Database.query_one(
            "SELECT * FROM grayscale_state WHERE release_id = ?", (release_id,)
        )
        if not state:
            return {"success": False, "message": "未找到灰度部署状态"}
        all_branches = [dict(r) for r in Database.query("SELECT * FROM branch_registry")]
        current_idx = state["phase_index"]
        if current_idx >= len(GRAYSCALE_PHASES):
            return {"success": False, "message": "灰度已完成，无法继续推进"}
        phase_enum, pct, hold_min = GRAYSCALE_PHASES[current_idx]
        count = max(1, len(all_branches) * pct // 100)
        deployed = random.sample(all_branches, min(count, len(all_branches)))
        deployed_codes = [b["code"] for b in deployed]
        next_idx = current_idx + 1
        is_final = next_idx >= len(GRAYSCALE_PHASES)
        new_phase = (
            GrayscalePhase.COMPLETED.value if is_final
            else GRAYSCALE_PHASES[next_idx][0].value
        )
        Database.execute(
            "UPDATE grayscale_state SET current_phase = ?, phase_index = ?, branches_deployed = ? WHERE release_id = ?",
            (new_phase, next_idx, json.dumps(deployed_codes), release_id),
        )
        self.audit.log(
            "GRAYSCALE_ADVANCE", "ReleaseApplication", release_id,
            json.dumps({"phase": phase_enum.value, "pct": pct, "branches": deployed_codes}, ensure_ascii=False),
        )
        if is_final:
            self._mark_stable(release_id)
            Database.execute(
                "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
                (ReleaseStatus.DEPLOYED.value, _now(), release_id),
            )
            self.audit.log(
                "GRAYSCALE_COMPLETED", "ReleaseApplication", release_id,
                "灰度发布完成，已标记为稳定版本",
            )
        return {
            "success": True,
            "phase": phase_enum.value,
            "pct": pct,
            "hold_minutes": hold_min,
            "branches_deployed": deployed_codes,
            "is_final": is_final,
            "next_phase": new_phase,
        }

    def get_state(self, release_id: str) -> Optional[Dict]:
        row = Database.query_one(
            "SELECT * FROM grayscale_state WHERE release_id = ?", (release_id,)
        )
        return dict(row) if row else None

    def _mark_stable(self, release_id: str):
        rel = Database.query_one(
            "SELECT * FROM release_applications WHERE id = ?", (release_id,)
        )
        if rel:
            sv_id = _gen_id("sv_")
            Database.execute(
                "INSERT INTO stable_versions (id, version, module, marked_at, release_id) VALUES (?, ?, ?, ?, ?)",
                (sv_id, rel["version"], rel["module"], _now(), release_id),
            )

    def _phase_info(self, release_id: str, idx: int, all_branches: list) -> Dict:
        if idx >= len(GRAYSCALE_PHASES):
            return {"release_id": release_id, "phase": "completed", "branches_deployed": []}
        phase_enum, pct, hold_min = GRAYSCALE_PHASES[idx]
        count = max(1, len(all_branches) * pct // 100)
        deployed = random.sample(all_branches, min(count, len(all_branches)))
        return {
            "release_id": release_id,
            "phase": phase_enum.value,
            "pct": pct,
            "hold_minutes": hold_min,
            "branches_deployed": [b["code"] for b in deployed],
        }


# ─── 6. MonitorEngine ────────────────────────────────────────────────────────

class MonitorEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger
        self._threads: Dict[str, threading.Thread] = {}
        self._stop_events: Dict[str, threading.Event] = {}
        self._lock = threading.Lock()
        self._callbacks: Dict[str, Callable] = {}

    def start(self, release_id: str, callback: Callable = None, auto_rollback: bool = False, mode: str = "thread"):
        with self._lock:
            if release_id in self._threads and self._threads[release_id].is_alive():
                logger.warning("Monitor already running for %s", release_id)
                return
            stop_event = threading.Event()
            self._stop_events[release_id] = stop_event
            if callback:
                self._callbacks[release_id] = callback
            t = threading.Thread(
                target=self._monitor_loop, args=(release_id, stop_event), daemon=True,
            )
            self._threads[release_id] = t
            t.start()

        run_id = _gen_id("monrun_")
        Database.execute(
            "INSERT INTO monitor_runs (id, release_id, started_at, auto_rollback_enabled, mode) VALUES (?, ?, ?, ?, ?)",
            (run_id, release_id, _now(), 1 if auto_rollback else 0, mode),
        )

        self.audit.log("MONITOR_START", "MonitorEngine", release_id, f"7x24监控已启动 (mode={mode}, auto_rollback={auto_rollback})")
        logger.info("Monitor started for release %s (mode=%s, auto_rollback=%s)", release_id, mode, auto_rollback)

    def stop(self, release_id: str):
        with self._lock:
            if release_id in self._stop_events:
                self._stop_events[release_id].set()
            self._callbacks.pop(release_id, None)

        last_run = Database.query_one(
            "SELECT * FROM monitor_runs WHERE release_id = ? ORDER BY started_at DESC LIMIT 1",
            (release_id,),
        )
        if last_run and not last_run["stopped_at"]:
            Database.execute(
                "UPDATE monitor_runs SET stopped_at = ? WHERE id = ?",
                (_now(), last_run["id"]),
            )

        self.audit.log("MONITOR_STOP", "MonitorEngine", release_id, "监控已停止")
        logger.info("Monitor stopped for release %s", release_id)

    def status(self, release_id: str) -> Dict:
        with self._lock:
            thread_running = (
                release_id in self._threads and self._threads[release_id].is_alive()
            )

        last_run = Database.query_one(
            "SELECT * FROM monitor_runs WHERE release_id = ? ORDER BY started_at DESC LIMIT 1",
            (release_id,),
        )

        last_snapshot = Database.query_one(
            "SELECT * FROM monitor_snapshots WHERE release_id = ? ORDER BY timestamp DESC LIMIT 1",
            (release_id,),
        )

        active = False
        if thread_running:
            active = True
        elif last_snapshot:
            try:
                snap_time = datetime.strptime(last_snapshot["timestamp"], "%Y-%m-%d %H:%M:%S")
                if (datetime.now() - snap_time).total_seconds() < 120:
                    active = True
            except (ValueError, TypeError):
                pass

        result = {
            "release_id": release_id,
            "monitoring": active,
            "thread_running": thread_running,
            "last_snapshot": dict(last_snapshot) if last_snapshot else None,
            "last_run": dict(last_run) if last_run else None,
            "thresholds": MONITOR_THRESHOLDS,
        }
        return result

    def run_foreground(self, release_id: str, callback: Callable = None, auto_rollback: bool = False) -> Dict:
        run_id = _gen_id("monrun_")
        Database.execute(
            "INSERT INTO monitor_runs (id, release_id, started_at, auto_rollback_enabled, mode) VALUES (?, ?, ?, ?, ?)",
            (run_id, release_id, _now(), 1 if auto_rollback else 0, "foreground"),
        )
        self.audit.log(
            "MONITOR_START", "MonitorEngine", release_id,
            f"前台监控已启动 (auto_rollback={auto_rollback})",
        )
        logger.info("Foreground monitor started for release %s", release_id)

        stop_event = threading.Event()
        if callback:
            with self._lock:
                self._callbacks[release_id] = callback

        try:
            while not stop_event.is_set():
                snapshot = self._collect_snapshot(release_id)
                self._save_snapshot(snapshot)

                status_str = "正常" if snapshot["status"] == MonitorStatus.RUNNING.value else "异常"
                print(
                    f"[{snapshot['timestamp']}] 监控状态: {status_str} | "
                    f"交易成功率: {snapshot['transaction_success_rate']}% | "
                    f"账务延迟: {snapshot['accounting_delay_ms']}ms | "
                    f"资金结算异常: {snapshot['fund_settlement_anomaly']}"
                )

                if snapshot["status"] == MonitorStatus.ANOMALY.value:
                    self.audit.log(
                        "MONITOR_ANOMALY", "MonitorEngine", release_id,
                        json.dumps(snapshot, ensure_ascii=False),
                    )
                    if callback:
                        try:
                            cb_result = callback(release_id, snapshot)
                            if auto_rollback:
                                print("[自动回退] 已触发合规回退，监控即将退出")
                                stop_event.set()
                                break
                        except Exception as e:
                            logger.error("Monitor callback error: %s", e)
                    elif auto_rollback:
                        print("[自动回退] 检测到异常但未配置回退回调")

                for _ in range(60):
                    if stop_event.is_set():
                        break
                    time.sleep(1)
        except KeyboardInterrupt:
            print("\n监控已停止 (用户中断)")
        finally:
            Database.execute(
                "UPDATE monitor_runs SET stopped_at = ? WHERE id = ?",
                (_now(), run_id),
            )
            with self._lock:
                self._callbacks.pop(release_id, None)
            self.audit.log("MONITOR_STOP", "MonitorEngine", release_id, "前台监控已停止")
            logger.info("Foreground monitor stopped for release %s", release_id)

        return {"release_id": release_id, "run_id": run_id, "mode": "foreground"}

    def _monitor_loop(self, release_id: str, stop_event: threading.Event):
        while not stop_event.is_set():
            snapshot = self._collect_snapshot(release_id)
            self._save_snapshot(snapshot)
            if snapshot["status"] == MonitorStatus.ANOMALY.value:
                self.audit.log(
                    "MONITOR_ANOMALY", "MonitorEngine", release_id,
                    json.dumps(snapshot, ensure_ascii=False),
                )
                with self._lock:
                    cb = self._callbacks.get(release_id)
                if cb:
                    try:
                        cb(release_id, snapshot)
                    except Exception as e:
                        logger.error("Monitor callback error: %s", e)
            stop_event.wait(60)

    def _collect_snapshot(self, release_id: str) -> Dict:
        simulate = os.environ.get("simulate_anomaly", "").lower() == "true"
        if simulate:
            success_rate = random.uniform(98.0, 99.5)
            delay_ms = random.uniform(600, 1500)
            anomaly = random.randint(1, 5)
        else:
            success_rate = random.uniform(99.92, 99.99)
            delay_ms = random.uniform(50, 400)
            anomaly = 0
        status = MonitorStatus.RUNNING.value
        if success_rate < MONITOR_THRESHOLDS["transaction_success_rate"]:
            status = MonitorStatus.ANOMALY.value
        if delay_ms > MONITOR_THRESHOLDS["accounting_delay_ms"]:
            status = MonitorStatus.ANOMALY.value
        if anomaly > MONITOR_THRESHOLDS["fund_settlement_anomaly"]:
            status = MonitorStatus.ANOMALY.value
        return {
            "id": _gen_id("mon_"),
            "release_id": release_id,
            "transaction_success_rate": round(success_rate, 4),
            "accounting_delay_ms": round(delay_ms, 2),
            "fund_settlement_anomaly": anomaly,
            "status": status,
            "timestamp": _now(),
        }

    def _save_snapshot(self, snapshot: Dict):
        Database.execute(
            "INSERT INTO monitor_snapshots (id, release_id, transaction_success_rate, accounting_delay_ms, fund_settlement_anomaly, status, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                snapshot["id"],
                snapshot["release_id"],
                snapshot["transaction_success_rate"],
                snapshot["accounting_delay_ms"],
                snapshot["fund_settlement_anomaly"],
                snapshot["status"],
                snapshot["timestamp"],
            ),
        )


# ─── 7. RollbackEngine ───────────────────────────────────────────────────────

class RollbackEngine:
    def __init__(self, audit_logger: AuditLogger, monitor_engine: MonitorEngine):
        self.audit = audit_logger
        self.monitor = monitor_engine

    def execute_rollback(self, release_id: str, reason: str, restart_monitor: bool = True) -> Dict:
        self.audit.log("ROLLBACK_START", "ReleaseApplication", release_id, f"原因: {reason}")

        # 1. stop monitoring
        self.monitor.stop(release_id)

        # 2. get release info
        rel_row = Database.query_one(
            "SELECT * FROM release_applications WHERE id = ?", (release_id,)
        )
        if not rel_row:
            return {"success": False, "message": "未找到发布申请"}
        rel = dict(rel_row)

        # 3. get previous stable version
        prev_version = ""
        if rel.get("previous_stable_version"):
            prev_version = rel["previous_stable_version"]
        else:
            prev_stable = self._get_previous_stable(rel["module"])
            if prev_stable:
                prev_version = prev_stable["version"]
        if not prev_version:
            return {"success": False, "message": "未找到可回退的稳定版本"}

        # 4. estimate affected accounts
        affected = self._estimate_affected_accounts(release_id)

        # 5. generate regulatory explanation
        reg_explanation = self._generate_regulatory_explanation(reason, affected)

        # 6. analyze root cause
        root_cause = self._analyze_root_cause(reason)

        # 7. notify stakeholders
        stakeholders = self._notify_stakeholders(release_id, reason)

        # 8. execute rollback / restore version
        self._execute_version_restore(release_id, prev_version)

        # 9. update release status
        Database.execute(
            "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
            (ReleaseStatus.ROLLED_BACK.value, _now(), release_id),
        )

        # 10. save rollback record
        rb_id = _gen_id("rb_")
        Database.execute(
            "INSERT INTO rollback_records (id, release_id, reason, previous_version, affected_accounts, regulatory_explanation, root_cause, status, created_at, notified_stakeholders) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                rb_id, release_id, reason, prev_version, affected,
                reg_explanation, root_cause, "completed", _now(),
                json.dumps(stakeholders, ensure_ascii=False),
            ),
        )

        # 11. restart 7x24 monitoring (use release_id for tracking)
        if restart_monitor:
            self.monitor.start(release_id)

        # 12. generate rollback report
        report = self._generate_rollback_report(
            rb_id, release_id, reason, prev_version,
            affected, reg_explanation, root_cause, stakeholders,
        )
        self.audit.log(
            "ROLLBACK_COMPLETE", "ReleaseApplication", release_id,
            json.dumps({
                "rollback_id": rb_id,
                "previous_version": prev_version,
                "affected": affected,
            }, ensure_ascii=False),
        )
        return {
            "success": True,
            "rollback_id": rb_id,
            "previous_version": prev_version,
            "affected_accounts": affected,
            "report": report,
        }

    def _get_previous_stable(self, module: str) -> Optional[Dict]:
        row = Database.query_one(
            "SELECT * FROM stable_versions WHERE module = ? ORDER BY marked_at DESC LIMIT 1",
            (module,),
        )
        return dict(row) if row else None

    def _estimate_affected_accounts(self, release_id: str) -> int:
        state = Database.query_one(
            "SELECT * FROM grayscale_state WHERE release_id = ?", (release_id,)
        )
        if state and state["branches_deployed"]:
            branches = json.loads(state["branches_deployed"])
            return len(branches) * random.randint(5000, 50000)
        return random.randint(10000, 100000)

    def _generate_regulatory_explanation(self, reason: str, affected: int) -> str:
        return (
            f"根据《银保监会银行业金融机构信息系统风险管理指引》及《银行业金融机构外包风险管理指引》相关要求，"
            f"本次回退操作原因：{reason}。影响账户数量约{affected}户。"
            f"我行已启动应急预案，确保客户资金安全，维护金融稳定。"
            f"回退操作符合监管要求，将在完成后续向银保监会提交书面报告。"
            f"依据：《关于进一步加强银行业金融机构信息安全管理的通知》（银监发〔2018〕4号）"
        )

    def _analyze_root_cause(self, reason: str) -> str:
        categories = [
            "代码缺陷：生产环境暴露未在测试环境发现的边界条件问题",
            "配置变更：环境配置差异导致运行时异常",
            "性能瓶颈：高并发场景下资源竞争导致处理延迟",
            "数据不一致：数据迁移过程中主备数据源同步异常",
            "第三方依赖：外部服务接口变更导致的兼容性问题",
        ]
        selected = random.choice(categories)
        return f"初步分析根因 - {selected}。详细原因：{reason}。需进一步排查确认。"

    def _notify_stakeholders(self, release_id: str, reason: str) -> List[Dict]:
        stakeholders = [
            {"role": "监管对接人", "name": "李监管", "notified_at": _now(), "channel": "电话+邮件"},
            {"role": "运维负责人", "name": "王运维", "notified_at": _now(), "channel": "电话+短信"},
            {"role": "业务负责人", "name": "张业务", "notified_at": _now(), "channel": "邮件+钉钉"},
            {"role": "合规负责人", "name": "赵合规", "notified_at": _now(), "channel": "邮件+电话"},
        ]
        for s in stakeholders:
            self.audit.log(
                "NOTIFY_STAKEHOLDER", "RollbackEngine", release_id,
                json.dumps(s, ensure_ascii=False),
            )
        return stakeholders

    def _execute_version_restore(self, release_id: str, version: str):
        self.audit.log(
            "VERSION_RESTORE", "ReleaseApplication", release_id,
            f"回退至稳定版本: {version}",
        )
        logger.info("Restored release %s to version %s", release_id, version)

    def _generate_rollback_report(
        self, rb_id, release_id, reason, prev_version,
        affected, reg_explanation, root_cause, stakeholders,
    ) -> Dict:
        report = {
            "rollback_id": rb_id,
            "release_id": release_id,
            "reason": reason,
            "restored_version": prev_version,
            "affected_accounts": affected,
            "regulatory_compliance": {
                "framework": "银保监会监管框架",
                "explanation": reg_explanation,
                "report_required": True,
                "report_deadline": (
                    datetime.now() + timedelta(hours=24)
                ).strftime("%Y-%m-%d %H:%M:%S"),
            },
            "root_cause_analysis": root_cause,
            "notified_stakeholders": stakeholders,
            "generated_at": _now(),
        }
        report_file = EXPORTS_DIR / f"rollback_report_{rb_id}.json"
        with open(report_file, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        return report


# ─── 8. DrillEngine ──────────────────────────────────────────────────────────

class DrillEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def create_drill(self, title: str, scenario: str) -> Dict:
        drill_id = _gen_id("drill_")
        plan = self._generate_plan(scenario)
        steps = self._generate_steps()
        Database.execute(
            "INSERT INTO drill_records (id, title, scenario, status, plan, steps, evidence, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                drill_id, title, scenario, DrillStatus.CREATED.value,
                json.dumps(plan, ensure_ascii=False),
                json.dumps(steps, ensure_ascii=False), "[]", _now(),
            ),
        )
        self.audit.log(
            "DRILL_CREATE", "DrillRecord", drill_id,
            json.dumps({"title": title, "scenario": scenario}, ensure_ascii=False),
        )
        return {"drill_id": drill_id, "title": title, "scenario": scenario, "plan": plan, "steps": steps}

    def execute_drill(self, drill_id: str) -> Dict:
        drill = Database.query_one(
            "SELECT * FROM drill_records WHERE id = ?", (drill_id,)
        )
        if not drill:
            return {"success": False, "message": "未找到演练记录"}
        steps = json.loads(drill["steps"])
        evidence_list: list = []
        for i, step in enumerate(steps):
            step["status"] = "executing"
            step["started_at"] = _now()
            evidence = self._collect_evidence(drill_id, step["name"], i + 1)
            step["status"] = "completed"
            step["completed_at"] = _now()
            step["evidence_id"] = evidence["id"]
            evidence_list.append(evidence)
            self.audit.log(
                "DRILL_STEP", "DrillRecord", drill_id,
                json.dumps({
                    "step": step["name"], "order": i + 1, "evidence": evidence["id"],
                }, ensure_ascii=False),
            )
        Database.execute(
            "UPDATE drill_records SET status = ?, steps = ?, evidence = ?, completed_at = ? WHERE id = ?",
            (
                DrillStatus.COMPLETED.value,
                json.dumps(steps, ensure_ascii=False),
                json.dumps(evidence_list, ensure_ascii=False),
                _now(), drill_id,
            ),
        )
        archive_path = self._archive_drill(drill_id, drill, steps, evidence_list)
        self.audit.log(
            "DRILL_COMPLETE", "DrillRecord", drill_id,
            json.dumps({"archive": str(archive_path)}, ensure_ascii=False),
        )
        return {
            "success": True,
            "drill_id": drill_id,
            "steps_completed": len(steps),
            "archive": str(archive_path),
        }

    def _generate_plan(self, scenario: str) -> List[str]:
        return [
            f"演练目标：验证{scenario}场景下的应急响应能力",
            "演练范围：核心交易系统、账务系统、支付清算系统",
            "演练参与方：总行科技部、总行业务部、总行合规部、运维团队",
            f"演练前提：{scenario}场景触发条件已满足",
            "预期结果：系统在规定时间内完成切换/恢复，业务连续性得到保障",
        ]

    def _generate_steps(self) -> List[Dict]:
        return [
            {
                "order": i + 1,
                "name": s,
                "status": "pending",
                "started_at": "",
                "completed_at": "",
                "evidence_id": "",
            }
            for i, s in enumerate(DRILL_STEPS)
        ]

    def _collect_evidence(self, drill_id: str, step_name: str, order: int) -> Dict:
        return {
            "id": _gen_id("ev_"),
            "drill_id": drill_id,
            "step_name": step_name,
            "order": order,
            "timestamp": _now(),
            "operator": "系统自动",
            "result": "通过",
            "details": f"步骤{order}({step_name})执行完成，验证通过",
        }

    def _archive_drill(self, drill_id: str, drill, steps, evidence) -> Path:
        archive_data = {
            "drill_id": drill_id,
            "title": drill["title"],
            "scenario": drill["scenario"],
            "status": DrillStatus.COMPLETED.value,
            "plan": json.loads(drill["plan"]),
            "steps": steps,
            "evidence": evidence,
            "created_at": drill["created_at"],
            "completed_at": _now(),
            "archived_at": _now(),
        }
        archive_path = ARCHIVES_DIR / f"drill_{drill_id}.json"
        with open(archive_path, "w", encoding="utf-8") as f:
            json.dump(archive_data, f, ensure_ascii=False, indent=2)
        return archive_path


# ─── 9. WeeklyReportEngine ───────────────────────────────────────────────────

class WeeklyReportEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def generate_weekly_stats(self, week_start: str = None) -> Dict:
        if not week_start:
            week_start = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        week_end = (
            datetime.strptime(week_start, "%Y-%m-%d") + timedelta(days=7)
        ).strftime("%Y-%m-%d")

        releases = Database.query(
            "SELECT * FROM release_applications WHERE created_at >= ? AND created_at < ?",
            (week_start, week_end),
        )
        release_list = [dict(r) for r in releases]
        total = len(release_list)
        success_count = sum(
            1 for r in release_list
            if r["status"] in [ReleaseStatus.DEPLOYED.value, ReleaseStatus.APPROVED.value]
        )
        rollback_count = sum(
            1 for r in release_list if r["status"] == ReleaseStatus.ROLLED_BACK.value
        )
        success_rate = (success_count / total * 100) if total > 0 else 0.0

        risk_dist: Dict[str, int] = {}
        for r in release_list:
            rl = r["risk_level"]
            risk_dist[rl] = risk_dist.get(rl, 0) + 1

        approval_nodes = Database.query(
            "SELECT * FROM approval_nodes WHERE approved_at >= ? AND approved_at < ? AND status = ?",
            (week_start, week_end, ApprovalStatus.APPROVED.value),
        )
        durations: list = []
        for node in approval_nodes:
            if node["approved_at"] and node["created_at"]:
                try:
                    created = datetime.strptime(node["created_at"], "%Y-%m-%d %H:%M:%S")
                    approved = datetime.strptime(node["approved_at"], "%Y-%m-%d %H:%M:%S")
                    durations.append((approved - created).total_seconds() / 60)
                except ValueError:
                    pass
        avg_duration = sum(durations) / len(durations) if durations else 0
        max_duration = max(durations) if durations else 0
        min_duration = min(durations) if durations else 0

        stats = {
            "week_start": week_start,
            "week_end": week_end,
            "release_count": total,
            "success_count": success_count,
            "success_rate": round(success_rate, 2),
            "rollback_count": rollback_count,
            "approval_duration_stats": {
                "avg_minutes": round(avg_duration, 2),
                "max_minutes": round(max_duration, 2),
                "min_minutes": round(min_duration, 2),
            },
            "risk_level_distribution": risk_dist,
        }
        self.audit.log(
            "WEEKLY_STATS", "WeeklyReportEngine", "",
            json.dumps(stats, ensure_ascii=False),
        )
        return stats

    def generate_pdf_report(self, stats: Dict) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas as pdf_canvas
            font_name = self._register_chinese_font()
            filename = EXPORTS_DIR / f"weekly_report_{stats['week_start']}.pdf"
            c = pdf_canvas.Canvas(str(filename), pagesize=A4)
            w, h = A4
            c.setFont(font_name, 16)
            c.drawString(20 * mm, h - 20 * mm, "银行核心系统发布周报")
            c.setFont(font_name, 10)
            y = h - 35 * mm
            lines = [
                f"报告周期：{stats['week_start']} 至 {stats['week_end']}",
                f"发布总数：{stats['release_count']}",
                f"成功数：{stats['success_count']}",
                f"成功率：{stats['success_rate']}%",
                f"回退次数：{stats['rollback_count']}",
                f"审批平均时长：{stats['approval_duration_stats']['avg_minutes']}分钟",
                f"审批最长时长：{stats['approval_duration_stats']['max_minutes']}分钟",
                f"审批最短时长：{stats['approval_duration_stats']['min_minutes']}分钟",
            ]
            for rl, cnt in stats["risk_level_distribution"].items():
                lines.append(f"风险级别 [{rl}]：{cnt}次")
            for line in lines:
                c.drawString(20 * mm, y, line)
                y -= 8 * mm
            c.save()
            self.audit.log("REPORT_PDF", "WeeklyReportEngine", "", str(filename))
            return str(filename)
        except ImportError:
            logger.warning("reportlab not available, falling back to text report")
            return self._fallback_text_report(stats)

    def generate_excel_report(self, stats: Dict) -> str:
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "周报统计"
            ws.append(["指标", "值"])
            data_rows = [
                ("报告周期", f"{stats['week_start']} 至 {stats['week_end']}"),
                ("发布总数", stats["release_count"]),
                ("成功数", stats["success_count"]),
                ("成功率", f"{stats['success_rate']}%"),
                ("回退次数", stats["rollback_count"]),
                ("审批平均时长(分钟)", stats["approval_duration_stats"]["avg_minutes"]),
                ("审批最长时长(分钟)", stats["approval_duration_stats"]["max_minutes"]),
                ("审批最短时长(分钟)", stats["approval_duration_stats"]["min_minutes"]),
            ]
            for row in data_rows:
                ws.append(row)
            ws2 = wb.create_sheet("风险级别分布")
            ws2.append(["风险级别", "次数"])
            for rl, cnt in stats["risk_level_distribution"].items():
                ws2.append([rl, cnt])
            filename = EXPORTS_DIR / f"weekly_report_{stats['week_start']}.xlsx"
            wb.save(str(filename))
            self.audit.log("REPORT_EXCEL", "WeeklyReportEngine", "", str(filename))
            return str(filename)
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV report")
            return self._fallback_csv_report(stats)

    def generate_trend_chart(self, stats_list: List[Dict]) -> str:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            font_path = self._find_chinese_font_path()
            if font_path:
                from matplotlib.font_manager import FontProperties, fontManager
                fontManager.addfont(font_path)
                font_prop = FontProperties(fname=font_path)
                font_name = font_prop.get_name()
                plt.rcParams["font.family"] = font_name
                plt.rcParams["axes.unicode_minus"] = False
            else:
                font_prop = None
            weeks = [s["week_start"] for s in stats_list]
            rates = [s["success_rate"] for s in stats_list]
            rollbacks = [s["rollback_count"] for s in stats_list]
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8))
            ax1.plot(weeks, rates, "b-o", label="Success Rate %")
            title1 = "周成功率趋势" if font_prop else "Weekly Success Rate Trend"
            ax1.set_title(title1)
            ax1.set_ylabel("%")
            ax1.legend()
            ax1.grid(True)
            ax2.bar(weeks, rollbacks, color="red", alpha=0.7, label="Rollback Count")
            title2 = "周回退次数趋势" if font_prop else "Weekly Rollback Count Trend"
            ax2.set_title(title2)
            ax2.set_ylabel("Count")
            ax2.legend()
            ax2.grid(True)
            plt.tight_layout()
            ts = _now().replace(":", "-").replace(" ", "_")
            filename = EXPORTS_DIR / f"trend_chart_{ts}.png"
            plt.savefig(str(filename), dpi=150)
            plt.close()
            self.audit.log("REPORT_CHART", "WeeklyReportEngine", "", str(filename))
            return str(filename)
        except ImportError:
            logger.warning("matplotlib not available, skipping chart generation")
            return ""

    def _register_chinese_font(self) -> str:
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                try:
                    from reportlab.pdfbase import pdfmetrics
                    from reportlab.pdfbase.ttfonts import TTFont
                    name = os.path.basename(fp).split(".")[0]
                    pdfmetrics.registerFont(TTFont(name, fp))
                    return name
                except Exception:
                    continue
        return "Helvetica"

    def _find_chinese_font_path(self) -> Optional[str]:
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        for fp in font_paths:
            if os.path.exists(fp):
                return fp
        return None

    def _fallback_text_report(self, stats: Dict) -> str:
        filename = EXPORTS_DIR / f"weekly_report_{stats['week_start']}.txt"
        lines = [
            "银行核心系统发布周报",
            "=" * 40,
            f"报告周期：{stats['week_start']} 至 {stats['week_end']}",
            f"发布总数：{stats['release_count']}",
            f"成功数：{stats['success_count']}",
            f"成功率：{stats['success_rate']}%",
            f"回退次数：{stats['rollback_count']}",
            f"审批平均时长：{stats['approval_duration_stats']['avg_minutes']}分钟",
            f"审批最长时长：{stats['approval_duration_stats']['max_minutes']}分钟",
            f"审批最短时长：{stats['approval_duration_stats']['min_minutes']}分钟",
        ]
        for rl, cnt in stats["risk_level_distribution"].items():
            lines.append(f"风险级别 [{rl}]：{cnt}次")
        with open(filename, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return str(filename)

    def _fallback_csv_report(self, stats: Dict) -> str:
        filename = EXPORTS_DIR / f"weekly_report_{stats['week_start']}.csv"
        with open(filename, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["指标", "值"])
            writer.writerow(["报告周期", f"{stats['week_start']} 至 {stats['week_end']}"])
            writer.writerow(["发布总数", stats["release_count"]])
            writer.writerow(["成功数", stats["success_count"]])
            writer.writerow(["成功率", f"{stats['success_rate']}%"])
            writer.writerow(["回退次数", stats["rollback_count"]])
            writer.writerow(["审批平均时长(分钟)", stats["approval_duration_stats"]["avg_minutes"]])
            writer.writerow(["审批最长时长(分钟)", stats["approval_duration_stats"]["max_minutes"]])
            writer.writerow(["审批最短时长(分钟)", stats["approval_duration_stats"]["min_minutes"]])
            for rl, cnt in stats["risk_level_distribution"].items():
                writer.writerow([f"风险级别[{rl}]", cnt])
        return str(filename)


# ─── 10. QueryEngine ─────────────────────────────────────────────────────────

class QueryEngine:
    def __init__(self, audit_logger: AuditLogger):
        self.audit = audit_logger

    def query_releases(
        self,
        start_time: str = None,
        end_time: str = None,
        version: str = None,
        module: str = None,
        status: str = None,
        branch_code: str = None,
        risk_level: str = None,
        applicant: str = None,
        limit: int = 100,
    ) -> List[Dict]:
        conditions = []
        params: list = []
        if start_time:
            conditions.append("created_at >= ?")
            params.append(start_time)
        if end_time:
            conditions.append("created_at <= ?")
            params.append(end_time)
        if version:
            conditions.append("version = ?")
            params.append(version)
        if module:
            conditions.append("module = ?")
            params.append(module)
        if status:
            conditions.append("status = ?")
            params.append(status)
        if branch_code:
            conditions.append("branch_code = ?")
            params.append(branch_code)
        if risk_level:
            conditions.append("risk_level = ?")
            params.append(risk_level)
        if applicant:
            conditions.append("applicant = ?")
            params.append(applicant)
        where = " AND ".join(conditions) if conditions else "1=1"
        sql = f"SELECT * FROM release_applications WHERE {where} ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        rows = Database.query(sql, tuple(params))
        return [dict(r) for r in rows]

    def batch_export(self, ids: List[str], fmt: str = "json") -> str:
        if not ids:
            return ""
        placeholders = ",".join("?" * len(ids))
        rows = Database.query(
            f"SELECT * FROM release_applications WHERE id IN ({placeholders})",
            tuple(ids),
        )
        data = [dict(r) for r in rows]
        ts = _now().replace(":", "-").replace(" ", "_")

        if fmt == "json":
            filename = EXPORTS_DIR / f"export_{ts}.json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        elif fmt == "csv":
            filename = EXPORTS_DIR / f"export_{ts}.csv"
            if data:
                with open(filename, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            else:
                with open(filename, "w", encoding="utf-8") as f:
                    f.write("")
        elif fmt == "zip":
            json_file = EXPORTS_DIR / f"export_{ts}_data.json"
            csv_file = EXPORTS_DIR / f"export_{ts}_data.csv"
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            if data:
                with open(csv_file, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            else:
                with open(csv_file, "w", encoding="utf-8") as f:
                    f.write("")
            zip_path = EXPORTS_DIR / f"export_{ts}.zip"
            with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(str(json_file), json_file.name)
                zf.write(str(csv_file), csv_file.name)
            json_file.unlink()
            csv_file.unlink(missing_ok=True)
            filename = zip_path
        else:
            filename = EXPORTS_DIR / f"export_{ts}.json"
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

        self.audit.log(
            "BATCH_EXPORT", "QueryEngine", "",
            json.dumps({"ids": ids, "format": fmt, "file": str(filename)}),
        )
        return str(filename)


# ─── 11. ReleaseOrchestrator ─────────────────────────────────────────────────

class ReleaseOrchestrator:
    def __init__(self):
        self.audit = AuditLogger()
        self.precheck = PreCheckEngine(self.audit)
        self.approval = ApprovalEngine(self.audit)
        self.grayscale = GrayscaleDeployEngine(self.audit)
        self.monitor = MonitorEngine(self.audit)
        self.rollback = RollbackEngine(self.audit, self.monitor)
        self.drill = DrillEngine(self.audit)
        self.report = WeeklyReportEngine(self.audit)
        self.query_engine = QueryEngine(self.audit)
        self._rollback_triggered = set()

    def submit_release(
        self, title: str, version: str, module: str, applicant: str,
        risk_level: str, branch_code: str = "", description: str = "",
    ) -> Dict:
        release_id = _gen_id("rel_")
        now = _now()

        prev_stable_row = Database.query_one(
            "SELECT * FROM stable_versions WHERE module = ? ORDER BY marked_at DESC LIMIT 1",
            (module,),
        )
        previous_stable_version = prev_stable_row["version"] if prev_stable_row else ""

        Database.execute(
            "INSERT INTO release_applications (id, title, version, module, applicant, risk_level, status, branch_code, description, previous_stable_version, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                release_id, title, version, module, applicant, risk_level,
                ReleaseStatus.SUBMITTED.value, branch_code, description,
                previous_stable_version, now, now,
            ),
        )
        self.audit.log(
            "SUBMIT", "ReleaseApplication", release_id,
            json.dumps({"title": title, "version": version, "risk_level": risk_level}, ensure_ascii=False),
        )

        precheck_result = self.precheck.run_checks(release_id, version, module)
        if not precheck_result["passed"]:
            Database.execute(
                "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
                (ReleaseStatus.PRECHECK_FAILED.value, _now(), release_id),
            )
            return {"success": False, "release_id": release_id, "precheck": precheck_result}

        Database.execute(
            "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
            (ReleaseStatus.PRECHECK_PASSED.value, _now(), release_id),
        )

        nodes = self.approval.create_workflow(release_id, risk_level)
        Database.execute(
            "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
            (ReleaseStatus.PENDING_APPROVAL.value, _now(), release_id),
        )
        return {
            "success": True,
            "release_id": release_id,
            "precheck": precheck_result,
            "approval_nodes": nodes,
        }

    def approve_release(self, release_id: str, role: str, approver: str,
                        comment: str = "", auto_deploy: bool = False) -> Dict:
        result = self.approval.approve_node(release_id, role, approver, comment)
        if result.get("all_approved"):
            Database.execute(
                "UPDATE release_applications SET status = ?, updated_at = ? WHERE id = ?",
                (ReleaseStatus.APPROVED.value, _now(), release_id),
            )
            self.audit.log(
                "FULLY_APPROVED", "ReleaseApplication", release_id, "所有审批节点已通过",
            )

            if auto_deploy:
                result["auto_deploy"] = self._try_auto_deploy(release_id)

        return result

    def _try_auto_deploy(self, release_id: str) -> Dict:
        try:
            init_result = self.start_grayscale_deployment(release_id)
            if not init_result.get("success"):
                fail_reason = init_result.get("message", "灰度初始化失败")
                Database.execute(
                    "UPDATE release_applications SET status = ?, deploy_failure_reason = ?, updated_at = ? WHERE id = ?",
                    (ReleaseStatus.APPROVED_DEPLOY_FAILED.value, fail_reason, _now(), release_id),
                )
                self.audit.log(
                    "AUTO_DEPLOY_FAILED", "ReleaseApplication", release_id,
                    json.dumps({"step": "灰度初始化", "reason": fail_reason}, ensure_ascii=False),
                )
                return {"success": False, "step": "灰度初始化", "reason": fail_reason}

            advance_result = self.advance_grayscale(release_id)
            if not advance_result.get("success"):
                fail_reason = advance_result.get("message", "第一阶段推送失败")
                Database.execute(
                    "UPDATE release_applications SET status = ?, deploy_failure_reason = ?, updated_at = ? WHERE id = ?",
                    (ReleaseStatus.APPROVED_DEPLOY_FAILED.value, fail_reason, _now(), release_id),
                )
                self.audit.log(
                    "AUTO_DEPLOY_FAILED", "ReleaseApplication", release_id,
                    json.dumps({"step": "第一阶段推送", "reason": fail_reason}, ensure_ascii=False),
                )
                return {"success": False, "step": "第一阶段推送", "reason": fail_reason}

            self.audit.log(
                "AUTO_DEPLOY_SUCCESS", "ReleaseApplication", release_id,
                json.dumps({
                    "phase": advance_result.get("phase", ""),
                    "pct": advance_result.get("pct", 0),
                }, ensure_ascii=False),
            )
            return {
                "success": True,
                "phase": advance_result.get("phase", ""),
                "pct": advance_result.get("pct", 0),
            }
        except Exception as e:
            fail_reason = f"自动部署异常: {str(e)}"
            Database.execute(
                "UPDATE release_applications SET status = ?, deploy_failure_reason = ?, updated_at = ? WHERE id = ?",
                (ReleaseStatus.APPROVED_DEPLOY_FAILED.value, fail_reason, _now(), release_id),
            )
            self.audit.log(
                "AUTO_DEPLOY_FAILED", "ReleaseApplication", release_id,
                json.dumps({"step": "未知", "reason": fail_reason, "traceback": traceback.format_exc()}, ensure_ascii=False),
            )
            return {"success": False, "step": "未知", "reason": fail_reason}

    def reject_release(self, release_id: str, role: str, approver: str,
                       comment: str = "") -> Dict:
        return self.approval.reject_node(release_id, role, approver, comment)

    def start_grayscale_deployment(self, release_id: str) -> Dict:
        rel = Database.query_one(
            "SELECT * FROM release_applications WHERE id = ?", (release_id,)
        )
        if not rel:
            return {"success": False, "message": "未找到发布申请"}
        if rel["status"] != ReleaseStatus.APPROVED.value:
            return {"success": False, "message": f"当前状态不允许部署: {rel['status']}"}
        init_result = self.grayscale.init_deployment(release_id)
        return {"success": True, "init": init_result}

    def advance_grayscale(self, release_id: str) -> Dict:
        result = self.grayscale.advance_phase(release_id)
        if result.get("is_final"):
            self.monitor.start(release_id, callback=self._on_monitor_anomaly)
        return result

    def _on_monitor_anomaly(self, release_id: str, snapshot: Dict):
        logger.warning("ANOMALY DETECTED for release %s: %s", release_id, snapshot)
        self.audit.log(
            "ANOMALY_CALLBACK", "ReleaseOrchestrator", release_id,
            json.dumps(snapshot, ensure_ascii=False),
        )

        if release_id in self._rollback_triggered:
            logger.warning("Rollback already triggered for release %s, skipping", release_id)
            return

        metrics_exceeded = []
        if snapshot["transaction_success_rate"] < MONITOR_THRESHOLDS["transaction_success_rate"]:
            metrics_exceeded.append(
                f"交易成功率={snapshot['transaction_success_rate']}% (阈值={MONITOR_THRESHOLDS['transaction_success_rate']}%)"
            )
        if snapshot["accounting_delay_ms"] > MONITOR_THRESHOLDS["accounting_delay_ms"]:
            metrics_exceeded.append(
                f"账务延迟={snapshot['accounting_delay_ms']}ms (阈值={MONITOR_THRESHOLDS['accounting_delay_ms']}ms)"
            )
        if snapshot["fund_settlement_anomaly"] > MONITOR_THRESHOLDS["fund_settlement_anomaly"]:
            metrics_exceeded.append(
                f"资金结算异常数={snapshot['fund_settlement_anomaly']} (阈值={MONITOR_THRESHOLDS['fund_settlement_anomaly']})"
            )

        reason = "监控异常触发合规回退 - " + "; ".join(metrics_exceeded) if metrics_exceeded else "监控指标异常"
        logger.warning("=" * 60)
        logger.warning("  AUTO-ROLLBACK TRIGGERED for release %s", release_id)
        logger.warning("  Reason: %s", reason)
        logger.warning("=" * 60)

        self._rollback_triggered.add(release_id)

        try:
            rb_result = self.rollback.execute_rollback(release_id, reason, restart_monitor=False)
            if rb_result["success"]:
                logger.warning(
                    "Auto-rollback completed. rollback_id=%s, restored_version=%s",
                    rb_result["rollback_id"], rb_result["previous_version"],
                )
                print(f"\n[自动回退] 回退ID: {rb_result['rollback_id']}")
                print(f"[自动回退] 触发指标: {'; '.join(metrics_exceeded) if metrics_exceeded else 'N/A'}")
                print(f"[自动回退] 恢复版本: {rb_result['previous_version']}")
            else:
                logger.error("Auto-rollback failed: %s", rb_result.get("message", "unknown"))
        except Exception as e:
            logger.error("Auto-rollback exception: %s", e)
            logger.error(traceback.format_exc())

    def trigger_rollback(self, release_id: str, reason: str) -> Dict:
        return self.rollback.execute_rollback(release_id, reason)

    def create_drill(self, title: str, scenario: str) -> Dict:
        return self.drill.create_drill(title, scenario)

    def execute_drill(self, drill_id: str) -> Dict:
        return self.drill.execute_drill(drill_id)

    def generate_weekly_report(self, week_start: str = None) -> Dict:
        stats = self.report.generate_weekly_stats(week_start)
        pdf_path = self.report.generate_pdf_report(stats)
        excel_path = self.report.generate_excel_report(stats)
        chart_path = self.report.generate_trend_chart([stats])
        return {"stats": stats, "pdf": pdf_path, "excel": excel_path, "chart": chart_path}

    def _get_last_week_range(self, day_of_week: int = 0) -> str:
        today = datetime.now()
        days_since_target = (today.weekday() - day_of_week) % 7
        if days_since_target == 0:
            days_since_target = 7
        last_target_day = today - timedelta(days=days_since_target)
        week_start = last_target_day.strftime("%Y-%m-%d")
        return week_start

    def _get_next_run_time(self, day_of_week: int = 0) -> datetime:
        today = datetime.now()
        days_until_target = (day_of_week - today.weekday()) % 7
        if days_until_target == 0:
            if today.hour >= 1:
                days_until_target = 7
        next_run = today + timedelta(days=days_until_target)
        next_run = next_run.replace(hour=0, minute=0, second=0, microsecond=0)
        return next_run

    def run_scheduler(self, mode: str = "foreground", day_of_week: int = 0) -> Dict:
        day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        day_name = day_names[day_of_week] if 0 <= day_of_week < 7 else "周一"

        if mode == "once":
            week_start = self._get_last_week_range(day_of_week)
            print(f"[调度器] 立即生成周报，周期起始: {week_start}")
            result = self.generate_weekly_report(week_start)
            self.audit.log(
                "SCHEDULER_REPORT", "Scheduler", "",
                json.dumps({
                    "mode": "once",
                    "week_start": week_start,
                    "pdf": result.get("pdf", ""),
                    "excel": result.get("excel", ""),
                    "chart": result.get("chart", ""),
                }, ensure_ascii=False),
            )
            return {"mode": "once", "week_start": week_start, "report": result}

        next_run = self._get_next_run_time(day_of_week)
        print(f"[调度器] 前台模式启动")
        print(f"[调度器] 报告日: 每周{day_name}")
        print(f"[调度器] 下次运行时间: {next_run.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"[调度器] 按 Ctrl+C 退出")

        last_generated_week = None

        try:
            while True:
                now = datetime.now()
                is_target_day = now.weekday() == day_of_week
                is_trigger_hour = 0 <= now.hour < 1

                week_start = self._get_last_week_range(day_of_week)

                if is_target_day and is_trigger_hour and last_generated_week != week_start:
                    print(f"\n[{now.strftime('%Y-%m-%d %H:%M:%S')}] 触发每周报告生成...")
                    try:
                        result = self.generate_weekly_report(week_start)
                        last_generated_week = week_start
                        print(f"  报告周期: {result['stats']['week_start']} 至 {result['stats']['week_end']}")
                        print(f"  PDF: {result.get('pdf', 'N/A')}")
                        print(f"  Excel: {result.get('excel', 'N/A')}")
                        print(f"  趋势图: {result.get('chart', 'N/A')}")
                        self.audit.log(
                            "SCHEDULER_REPORT", "Scheduler", "",
                            json.dumps({
                                "mode": "foreground",
                                "week_start": week_start,
                                "pdf": result.get("pdf", ""),
                                "excel": result.get("excel", ""),
                                "chart": result.get("chart", ""),
                            }, ensure_ascii=False),
                        )
                        next_run = self._get_next_run_time(day_of_week)
                        print(f"  下次运行时间: {next_run.strftime('%Y-%m-%d %H:%M:%S')}")
                    except Exception as e:
                        logger.error("Weekly report generation failed: %s", e)
                        logger.error(traceback.format_exc())

                time.sleep(3600)
        except KeyboardInterrupt:
            print("\n[调度器] 已停止")
            return {"mode": "foreground", "stopped": True}

        return {"mode": "foreground", "stopped": True}

    def generate_audit_excel(self) -> str:
        rows = Database.query("SELECT * FROM audit_log_chain ORDER BY id ASC")
        data = [dict(r) for r in rows]
        ts = _now().replace(":", "-").replace(" ", "_")
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "审计日志"
            if data:
                ws.append(list(data[0].keys()))
                for row in data:
                    ws.append(list(row.values()))
            filename = EXPORTS_DIR / f"audit_export_{ts}.xlsx"
            wb.save(str(filename))
            return str(filename)
        except ImportError:
            filename = EXPORTS_DIR / f"audit_export_{ts}.csv"
            with open(filename, "w", encoding="utf-8-sig", newline="") as f:
                if data:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            return str(filename)

    def query_history(self, **filters) -> List[Dict]:
        return self.query_engine.query_releases(**filters)

    def batch_export(self, ids: List[str], fmt: str = "json") -> str:
        return self.query_engine.batch_export(ids, fmt)

    def verify_audit_integrity(self) -> Dict:
        return self.audit.verify_integrity()

    def monitor_action(self, release_id: str, action: str, auto_rollback: bool = False) -> Dict:
        if action == "start":
            self.monitor.start(release_id, callback=self._on_monitor_anomaly, auto_rollback=auto_rollback, mode="thread")
            return {"success": True, "action": "start", "release_id": release_id, "auto_rollback": auto_rollback}
        elif action == "stop":
            self.monitor.stop(release_id)
            return {"success": True, "action": "stop", "release_id": release_id}
        elif action == "status":
            return self.monitor.status(release_id)
        elif action == "run":
            return self.monitor.run_foreground(
                release_id,
                callback=self._on_monitor_anomaly,
                auto_rollback=auto_rollback,
            )
        else:
            return {"success": False, "message": f"未知监控操作: {action}"}

    def get_release(self, release_id: str) -> Optional[Dict]:
        row = Database.query_one(
            "SELECT * FROM release_applications WHERE id = ?", (release_id,)
        )
        return dict(row) if row else None


# ─── Demo ─────────────────────────────────────────────────────────────────────

def run_demo():
    print("=" * 60)
    print("  银行核心系统发布与合规回退自动化管理 - 全流程演示")
    print("=" * 60)

    init_db()
    orch = ReleaseOrchestrator()

    print("\n[1] 提交发布申请...")
    result = orch.submit_release(
        title="核心账务系统v4.1.0升级",
        version="4.1.0",
        module="core",
        applicant="张工",
        risk_level=RiskLevel.NIGHTTIME_BATCH.value,
        branch_code="BJ001",
        description="核心账务模块功能优化及性能提升",
    )
    release_id = result["release_id"]
    print(f"  发布ID: {release_id}")
    print(f"  预检结果: {'通过' if result['success'] else '未通过'}")
    if result["success"]:
        print(f"  审批节点: {[n['role'] for n in result['approval_nodes']]}")

    if result["success"]:
        rel_info = orch.get_release(release_id)
        print(f"  前一稳定版本: {rel_info.get('previous_stable_version', 'N/A')}")

        print("\n[2] 审批流程 (最后一个节点自动部署)...")
        nodes = result["approval_nodes"]
        for i, node in enumerate(nodes):
            is_last = i == len(nodes) - 1
            apr = orch.approve_release(
                release_id, node["role"], f"{node['role']}负责人", "同意发布",
                auto_deploy=is_last,
            )
            print(f"  {node['role']}: {'已批准' if apr['success'] else '审批失败'}")
            if apr.get("all_approved"):
                print("  ✓ 全部审批通过")
                if apr.get("auto_deploy"):
                    ad = apr["auto_deploy"]
                    if ad.get("success"):
                        print(f"  ✓ 自动部署成功: 阶段={ad.get('phase')}, 比例={ad.get('pct')}%")
                    else:
                        print(f"  ✗ 自动部署失败: 步骤={ad.get('step')}, 原因={ad.get('reason')}")

        print("\n[3] 灰度发布 (继续推进)...")
        rel = orch.get_release(release_id)
        current_status = rel.get("status", "")
        phase_names = ["20%", "50%", "100%"]
        if current_status in [ReleaseStatus.DEPLOYING.value, ReleaseStatus.APPROVED_DEPLOY_FAILED.value]:
            if current_status == ReleaseStatus.APPROVED_DEPLOY_FAILED.value:
                print(f"  当前状态: 审批通过但部署失败 ({rel.get('deploy_failure_reason', '')})")
                print("  重新初始化部署...")
                orch.start_grayscale_deployment(release_id)
            for name in phase_names:
                adv = orch.advance_grayscale(release_id)
                branches_count = len(adv.get("branches_deployed", []))
                is_final = adv.get("is_final", False)
                print(f"  阶段 {name}: branches={branches_count}, 完成={is_final}")
                if is_final:
                    print("  ✓ 灰度发布完成")
        else:
            deploy = orch.start_grayscale_deployment(release_id)
            print(f"  灰度初始化: {'成功' if deploy['success'] else '失败'}")
            phase_names_all = ["5%", "20%", "50%", "100%"]
            for name in phase_names_all:
                adv = orch.advance_grayscale(release_id)
                branches_count = len(adv.get("branches_deployed", []))
                is_final = adv.get("is_final", False)
                print(f"  阶段 {name}: branches={branches_count}, 完成={is_final}")
                if is_final:
                    print("  ✓ 灰度发布完成")

        print("\n[4] 监控状态...")
        time.sleep(2)
        ms = orch.monitor_action(release_id, "status")
        print(f"  监控运行中: {ms.get('monitoring', False)}")
        print(f"  线程运行中: {ms.get('thread_running', False)}")
        if ms.get("last_snapshot"):
            snap = ms["last_snapshot"]
            print(f"  最近快照: 成功率={snap['transaction_success_rate']}%, 延迟={snap['accounting_delay_ms']}ms")
        if ms.get("last_run"):
            run = ms["last_run"]
            print(f"  最近运行: 启动于={run.get('started_at', 'N/A')}, 模式={run.get('mode', 'N/A')}")
        print(f"  阈值配置: {ms.get('thresholds', {})}")

        print("\n[5] 模拟异常回退 (验证回退到前一稳定版本)...")
        rel_before = orch.get_release(release_id)
        print(f"  当前版本: {rel_before['version']}")
        print(f"  前一稳定版本 (记录): {rel_before.get('previous_stable_version', 'N/A')}")
        rb = orch.trigger_rollback(release_id, "交易成功率低于阈值，触发合规回退")
        print(f"  回退结果: {'成功' if rb['success'] else '失败'}")
        if rb["success"]:
            print(f"  回退ID: {rb['rollback_id']}")
            print(f"  恢复版本: {rb['previous_version']}")
            print(f"  影响账户: {rb['affected_accounts']}")
            print(f"  报告中的恢复版本: {rb['report']['restored_version']}")
            assert rb["previous_version"] == rb["report"]["restored_version"], "回退版本不一致!"
            print(f"  ✓ 回退版本与报告版本一致")
            assert rb["previous_version"] == rel_before.get("previous_stable_version", ""), "回退版本与记录的前一稳定版本不一致!"
            print(f"  ✓ 回退版本与记录的前一稳定版本一致")
            explanation = rb["report"]["regulatory_compliance"]["explanation"][:80]
            print(f"  监管说明: {explanation}...")

    print("\n[6] 演练管理...")
    drill = orch.create_drill("核心系统灾备切换演练", "主数据中心故障切换")
    drill_id = drill["drill_id"]
    print(f"  演练ID: {drill_id}")
    print(f"  演练步骤: {[s['name'] for s in drill['steps']]}")
    drill_exec = orch.execute_drill(drill_id)
    print(f"  演练执行: {'成功' if drill_exec['success'] else '失败'}")
    if drill_exec["success"]:
        print(f"  完成步骤数: {drill_exec['steps_completed']}")
        print(f"  归档路径: {drill_exec['archive']}")

    print("\n[7] 周报生成...")
    report = orch.generate_weekly_report()
    print(f"  发布总数: {report['stats']['release_count']}")
    print(f"  成功率: {report['stats']['success_rate']}%")
    print(f"  PDF报告: {report['pdf']}")
    print(f"  Excel报告: {report['excel']}")

    print("\n[8] 审计完整性验证...")
    integrity = orch.verify_audit_integrity()
    print(f"  完整性: {'通过' if integrity['valid'] else '失败'}")
    print(f"  审计条目: {integrity['entries']}")

    print("\n[9] 查询与导出...")
    releases = orch.query_history(module="core", limit=10)
    print(f"  查询结果: {len(releases)}条记录")
    if releases:
        exp = orch.batch_export([r["id"] for r in releases], "zip")
        print(f"  导出路径: {exp}")

    print("\n[10] 调度器 (立即生成模式演示)...")
    sched_result = orch.run_scheduler(mode="once", day_of_week=0)
    if sched_result.get("report"):
        rpt = sched_result["report"]
        print(f"  周起始: {sched_result.get('week_start', 'N/A')}")
        print(f"  发布数: {rpt['stats']['release_count']}")
        print(f"  成功率: {rpt['stats']['success_rate']}%")
        print(f"  PDF报告: {rpt.get('pdf', 'N/A')}")
        print(f"  Excel报告: {rpt.get('excel', 'N/A')}")
        print(f"  趋势图: {rpt.get('chart', 'N/A')}")

    print("\n" + "=" * 60)
    print("  全流程演示完成!")
    print()
    print("  新增功能:")
    print("  1. 监控异常自动触发合规回退")
    print("  2. 回退恢复到前一稳定版本 (而非当前版本)")
    print("  3. 前台监控模式 (monitor --action run --auto-rollback)")
    print("  4. 审批通过后自动部署 (approve --auto-deploy)")
    print("  5. 周报调度器 (scheduler --mode foreground/once --day 0)")
    print("=" * 60)


# ─── 12. CLI main() ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="银行核心系统发布与合规回退自动化管理",
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    sub_submit = subparsers.add_parser("submit", help="提交发布申请")
    sub_submit.add_argument("--title", required=True, help="发布标题")
    sub_submit.add_argument("--version", required=True, help="版本号")
    sub_submit.add_argument("--module", required=True, help="模块名")
    sub_submit.add_argument("--applicant", required=True, help="申请人")
    sub_submit.add_argument(
        "--risk-level", required=True,
        choices=[rl.value for rl in RiskLevel], help="风险级别",
    )
    sub_submit.add_argument("--branch-code", default="", help="分支代码")
    sub_submit.add_argument("--description", default="", help="描述")

    sub_approve = subparsers.add_parser("approve", help="审批发布")
    sub_approve.add_argument("--release-id", required=True, help="发布ID")
    sub_approve.add_argument("--role", required=True, help="审批角色")
    sub_approve.add_argument("--approver", required=True, help="审批人")
    sub_approve.add_argument("--comment", default="同意", help="审批意见")
    sub_approve.add_argument("--auto-deploy", action="store_true", help="审批通过后自动部署")

    sub_reject = subparsers.add_parser("reject", help="拒绝发布")
    sub_reject.add_argument("--release-id", required=True, help="发布ID")
    sub_reject.add_argument("--role", required=True, help="审批角色")
    sub_reject.add_argument("--approver", required=True, help="审批人")
    sub_reject.add_argument("--comment", default="拒绝", help="拒绝原因")

    sub_deploy = subparsers.add_parser("deploy", help="灰度部署推进")
    sub_deploy.add_argument("--release-id", required=True, help="发布ID")

    sub_rollback = subparsers.add_parser("rollback", help="触发回退")
    sub_rollback.add_argument("--release-id", required=True, help="发布ID")
    sub_rollback.add_argument("--reason", required=True, help="回退原因")

    sub_monitor = subparsers.add_parser("monitor", help="监控管理")
    sub_monitor.add_argument("--release-id", required=True, help="发布ID")
    sub_monitor.add_argument("--action", required=True, choices=["start", "stop", "status", "run"], help="操作")
    sub_monitor.add_argument("--auto-rollback", action="store_true", help="检测到异常时自动回退")

    sub_drill = subparsers.add_parser("drill", help="演练管理")
    sub_drill.add_argument("--action", required=True, choices=["create", "execute"], help="操作")
    sub_drill.add_argument("--title", default="", help="演练标题")
    sub_drill.add_argument("--scenario", default="", help="演练场景")
    sub_drill.add_argument("--drill-id", default="", help="演练ID")

    sub_report = subparsers.add_parser("report", help="报告生成")
    sub_report.add_argument("--type", required=True, choices=["weekly", "pdf", "excel"], help="报告类型")
    sub_report.add_argument("--week-start", default=None, help="周起始日期(YYYY-MM-DD)")

    sub_query = subparsers.add_parser("query", help="查询发布记录")
    sub_query.add_argument("--filters", default="{}", help="查询条件(JSON)")

    sub_export = subparsers.add_parser("export", help="批量导出")
    sub_export.add_argument("--ids", required=True, help="发布ID(逗号分隔)")
    sub_export.add_argument("--format", default="json", choices=["json", "csv", "zip"], help="导出格式")

    sub_audit = subparsers.add_parser("audit", help="审计管理")
    sub_audit.add_argument("--action", required=True, choices=["verify", "query"], help="操作")

    sub_demo = subparsers.add_parser("demo", help="运行全流程演示")

    sub_scheduler = subparsers.add_parser("scheduler", help="周报调度器")
    sub_scheduler.add_argument("--mode", default="foreground", choices=["foreground", "once"], help="运行模式")
    sub_scheduler.add_argument("--day", type=int, default=0, help="每周几生成 (0=周一, 1=周二, ..., 6=周日)")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    init_db()
    orch = ReleaseOrchestrator()

    if args.command == "submit":
        result = orch.submit_release(
            title=args.title, version=args.version, module=args.module,
            applicant=args.applicant, risk_level=args.risk_level,
            branch_code=args.branch_code, description=args.description,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "approve":
        result = orch.approve_release(
            args.release_id, args.role, args.approver, args.comment,
            auto_deploy=args.auto_deploy,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "reject":
        result = orch.reject_release(
            args.release_id, args.role, args.approver, args.comment,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "deploy":
        rel = orch.get_release(args.release_id)
        if not rel:
            print(json.dumps({"error": "未找到发布申请"}, ensure_ascii=False))
            return
        if rel["status"] in [ReleaseStatus.APPROVED.value, ReleaseStatus.APPROVED_DEPLOY_FAILED.value]:
            orch.start_grayscale_deployment(args.release_id)
        result = orch.advance_grayscale(args.release_id)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "rollback":
        result = orch.trigger_rollback(args.release_id, args.reason)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "monitor":
        result = orch.monitor_action(args.release_id, args.action, auto_rollback=args.auto_rollback)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "drill":
        if args.action == "create":
            result = orch.create_drill(args.title, args.scenario)
        elif args.action == "execute":
            result = orch.execute_drill(args.drill_id)
        else:
            result = {"error": "未知演练操作"}
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "report":
        result = orch.generate_weekly_report(args.week_start)
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "query":
        filters = json.loads(args.filters)
        results = orch.query_history(**filters)
        print(json.dumps(results, ensure_ascii=False, indent=2, default=str))

    elif args.command == "export":
        ids = [i.strip() for i in args.ids.split(",") if i.strip()]
        filepath = orch.batch_export(ids, args.format)
        print(json.dumps({"export_path": filepath}, ensure_ascii=False))

    elif args.command == "audit":
        if args.action == "verify":
            result = orch.verify_audit_integrity()
        elif args.action == "query":
            result = orch.audit.query(limit=50)
        else:
            result = {"error": "未知审计操作"}
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))

    elif args.command == "demo":
        run_demo()

    elif args.command == "scheduler":
        result = orch.run_scheduler(mode=args.mode, day_of_week=args.day)
        if args.mode == "once":
            print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()